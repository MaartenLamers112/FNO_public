"""Read-only vergelijkrapport tussen FNO en Maior Memorix."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.enums.comparison_status import ComparisonStatus
from app.exceptions import NotFoundError
from app.repositories import PersonRepository, PhotoRepository
from app.services.comparison_service import ComparisonService
from app.services.memorix_description_parser import MemorixDescriptionParser
from app.services.memorix_service import MemorixService


@dataclass(frozen=True)
class MmComparisonReportRow:
    """Eén verschil waarvoor handmatige controle nodig is."""

    photo_number: str
    mm_id: str
    action: str
    field: str
    fno_value: str
    mm_value: str
    mm_url: str


class MmComparisonReportService:
    """Bouw een read-only werkset met verschillen tussen FNO en MM."""

    FIELD_LABELS = {
        "subject": "Onderwerp",
        "date": "Datering",
        "location": "Plaats",
        "description": "Beschrijving",
    }

    def __init__(
        self,
        *,
        photo_repository: PhotoRepository | None = None,
        person_repository: PersonRepository | None = None,
        memorix_service: MemorixService | None = None,
        comparison_service: ComparisonService | None = None,
        description_parser: MemorixDescriptionParser | None = None,
    ) -> None:
        """Initialiseer de service met leesafhankelijkheden."""

        self.photo_repository = photo_repository or PhotoRepository()
        self.person_repository = person_repository or PersonRepository()
        self.memorix_service = memorix_service or MemorixService()
        self.comparison_service = comparison_service or ComparisonService()
        self.description_parser = description_parser or MemorixDescriptionParser()

    def build_rows(self) -> list[MmComparisonReportRow]:
        """Vergelijk alle geïmporteerde foto's en geef alleen actiepunten terug."""

        rows: list[MmComparisonReportRow] = []
        for photo in self.photo_repository.get_all():
            try:
                fields, metadata = self.memorix_service.get_metadata_record(photo.mm_id)
            except NotFoundError:
                rows.append(
                    MmComparisonReportRow(
                        photo_number=photo.photo_number,
                        mm_id=photo.mm_id,
                        action="MM-record ontbreekt",
                        field="Record",
                        fno_value=photo.photo_number,
                        mm_value="Niet gevonden",
                        mm_url="",
                    )
                )
                continue

            parsed = self.description_parser.parse(metadata.description_raw)
            comparison = self.comparison_service.compare_photo(
                mm_data={
                    **metadata.as_local_values(),
                    "description": parsed.description,
                },
                fno_data={
                    "subject": photo.local_subject or "",
                    "date": photo.local_date or "",
                    "location": photo.local_location or "",
                    "description": photo.local_description or "",
                },
                mm_names=parsed.names,
                fno_persons=self.person_repository.get_by_photo(photo.id),
                names_reliable=parsed.reliable,
                reason=parsed.reason,
            )
            mm_url = self._get_mm_url(fields)

            for field in comparison.fields:
                if field.status == ComparisonStatus.GREEN:
                    continue
                rows.append(
                    MmComparisonReportRow(
                        photo_number=photo.photo_number,
                        mm_id=photo.mm_id,
                        action=(
                            "Handmatig controleren"
                            if field.status == ComparisonStatus.RED
                            else "Verschil controleren"
                        ),
                        field=self.FIELD_LABELS[field.field],
                        fno_value=field.fno_value,
                        mm_value=field.mm_value,
                        mm_url=mm_url,
                    )
                )

            for person in comparison.persons:
                if person.status == ComparisonStatus.GREEN:
                    continue
                rows.append(
                    MmComparisonReportRow(
                        photo_number=photo.photo_number,
                        mm_id=photo.mm_id,
                        action=(
                            "Handmatig controleren"
                            if person.status == ComparisonStatus.RED
                            else "Verschil controleren"
                        ),
                        field=f"Persoon {person.label_number}",
                        fno_value=person.fno_name,
                        mm_value=person.mm_name,
                        mm_url=mm_url,
                    )
                )

            if photo.person_display_mode == "numbered":
                rows.extend(self._check_numbered_photo(photo.photo_number, photo.mm_id))

        return rows

    def build_xlsx(self) -> bytes:
        """Maak het vergelijkrapport als Excel-bestand in geheugen."""

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Actiepunten"
        headers = (
            "Fotonummer",
            "MM ID",
            "Actie",
            "Veld",
            "FNO-waarde",
            "MM-waarde",
            "MM-link",
        )
        sheet.append(headers)

        for row in self.build_rows():
            sheet.append((
                row.photo_number,
                row.mm_id,
                row.action,
                row.field,
                row.fno_value,
                row.mm_value,
                row.mm_url,
            ))
            if row.mm_url:
                link_cell = sheet.cell(row=sheet.max_row, column=7)
                link_cell.hyperlink = row.mm_url
                link_cell.style = "Hyperlink"

        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = (16, 28, 24, 20, 42, 42, 44)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _check_numbered_photo(
        self,
        photo_number: str,
        mm_id: str,
    ) -> list[MmComparisonReportRow]:
        """Controleer of de bijbehorende n-foto voor numbered-mode in MM bestaat."""

        if photo_number.casefold().endswith("n"):
            return []

        expected_number = f"{photo_number}n"
        records = self.memorix_service.search_records(
            filters={"dc_identifier": expected_number},
            rows=10,
        )
        matches = [
            self.memorix_service.normalize_search_record(record)
            for record in records
        ]
        exact_matches = [
            record
            for record in matches
            if str(record.get("photo_number", "")).casefold()
            == expected_number.casefold()
        ]
        if exact_matches:
            return []

        return [
            MmComparisonReportRow(
                photo_number=photo_number,
                mm_id=mm_id,
                action="Numbered-mode foto ontbreekt in MM",
                field="MM-foto",
                fno_value=expected_number,
                mm_value="Niet gevonden",
                mm_url="",
            )
        ]

    @staticmethod
    def _get_mm_url(fields: dict[str, Any]) -> str:
        """Geef de publieke recordlink uit MM terug wanneer beschikbaar."""

        for key in (
            "delving_landingPage",
            "europeana_isShownAt",
            "edm_isShownAt",
            "delving_uri",
        ):
            value = fields.get(key)
            if isinstance(value, list):
                value = next((item for item in value if str(item).strip()), "")
            if isinstance(value, str) and value.startswith(("https://", "http://")):
                return value.strip()
        return ""
